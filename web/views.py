#from msilib.schema import Billboard
from django.shortcuts import render,redirect
from .models import Document
from .forms import PostForm,UploadFileForm,handle_uploaded_file
from django.http import HttpResponse,HttpResponseRedirect 
from .models import Candidate
import csv
import pandas as pd

"""
# Create your views here.
def test(request):
    if request.method=='GET': 
        modelform=PostForm()
        return render(request,'webtest.html',{'modelform':modelform}) #postform을 변수에 담아서 webtest.html 을 렌더링
#템플릿 불러오기 
    elif request.method=='POST':
        modelform=PostForm(request.POST) #입력된 내용들 변수에 저장
        if modelform.is_valid(): #form 이 유효하면 
            post=modelform.save(commit=False) #데이터가져옴
            post.files=request.FILES('file')
            post.save() #DB에 저장
            return redirect('webtest.html'+str(post.id)) #URL로 이동

def read(requests,bid):
    post=Document.objects.prefetch_related('post_set').get(id=bid)
    return render(requests,'web/list.html',{'post':post})
"""


def upload_file(request):
    if request.method == 'POST':
       
        form = UploadFileForm(request.POST, request.FILES)
        print(request.FILES)
        print(request.POST)
        print(form)
        if form.is_valid():
            handle_uploaded_file(request.FILES['file'])
            return HttpResponseRedirect('webtest')
    else:
        form = UploadFileForm()

    return render(request, 'templatemo_555_upright/.html', {'form': form})
def main(request):
    return render(request, 'html/index.html')
"""
def csvTomodel(request):
    path='/Users/songryu/Desktop/capstonedesign_backend/web/media/files/data.xlsx'
    file=open(path)
    reader=csv.reader(file)
    print('----',reader)
    list=[]
    for row in reader:
        list.append(seops(a=row[0],
                          b=row[1],
                          c=row[2]))
    seops.objects.bulk_create(list)                    
    return HttpResponse('create model --')


# 엑셀을 생성 및 행 추가 (여기서는 행 단위 추가만 함 열만 추가하는건 검색 바람 )
from openpyxl import Workbook # 엑셀을 만드는 api (엑셀 미설치 시에도 동작)
from io import BytesIO # 엑셀 파일을 전송 할 수 있도록 바이트 배열로 변환

wb = Workbook()  # 엑셀 생성
ws = wb.active	# 엑셀 활성화
excelfile = BytesIO() #바이트 배열 생성

ws['A1']= 'company' # 엑셀 a1 열 이름 정하기
ws['B1']= 'product'
ws['C1']= 'count'

for i in text: # text 는 db 에 저장된 내용 전체
  content=[i.company,i.product_name,i.count]   #리스트 형태로 1 행씩 생성(a1, b1, c1) 에 각각
  ws.append(content) # 엑셀에 1행을 추가

wb.close()  #엑셀 닫기
wb.save(excelfile) # 바이트배열로 저장 (mail 전송 하려면 바이트형태로 변환 되어야 함)
"""
def main_view(request):
    with open('web/media/files/data.xlsx','r') as f:
        dr = csv.DictReader(f)
        s = pd.DataFrame(dr)
    ss = []
    for i in range(len(s)):
        st = (s['이름'][i], s['나이'][i], s['주소'][i])
        ss.append(st)
    for i in range(len(s)):
        Candidate.objects.create(name=ss[i][0], code=ss[i][1], ipo_date=ss[i][2])
        context={'df':s.to_html(justify='center')}
        return render(request,'classes.html',context)
